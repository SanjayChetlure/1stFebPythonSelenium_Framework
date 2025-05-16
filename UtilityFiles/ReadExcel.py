import openpyxl

class ReadTD:

    @staticmethod
    def getTestData(rowIndex,colIndex):
        workbook = openpyxl.load_workbook(
            "D:\\Other_Notes\\Python\\Workspace\\1stFebPythonSelenium_Framework\\TestData\\Data.xlsx")
        sheet = workbook['Sheet2']

        s2 = sheet.cell(row=rowIndex, column=colIndex).value
        return s2

    @staticmethod
    def getTestDataInt(rowIndex,colIndex):
        workbook = openpyxl.load_workbook(
            "D:\\Other_Notes\\Python\\Workspace\\1stFebPythonSelenium_Framework\\TestData\\Data.xlsx")
        sheet = workbook['Sheet2']

        s2 = sheet.cell(row=rowIndex, column=colIndex).value
        s2=int(s2)          #convert String to int
        return s2

